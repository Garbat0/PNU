# -*- coding: utf-8 -*-\
"""
Created on Fri Feb 18 16:25:37 2022

@author: gellison
"""
from openpyxl import Workbook
#from openpyxl.utils import get_column_letter
import pandas as pd
import ast
import re
import glob

wb = Workbook()

dest_filename = 'Synchronous_Dashboard.xlsx'

ws1 = wb.active
ws1.title = 'Meta'

#for row in range(1,40):
#        ws1.append(range(600))
#        
#ws2 = wb.create_sheet(title = "pi")
#
#ws2['F5'] = 3.14
#
#ws3 = wb.create_sheet(title="data")
#for row in range(1,10):
#    for col in range(1,27):
#        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
#print(ws3['A1'].value)

trunknlist= ['']
limblist = ['Limb1']
templist = [[f'=HYPERLINK("C:/Users/gellison/OneDrive - FEMA/Desktop/LocalCopy/PersonalNotes/Python Resources/Programs/Root/Trunk1/Limb 11/Bough 11", \"Bough 11\")',
            f'=HYPERLINK("C:/Users/gellison/OneDrive - FEMA/Desktop/LocalCopy/PersonalNotes/Python Resources/Programs/Root/Trunk1/Limb 11/Bough 11/Branch 1", \"Branch 1\")',
            f'=HYPERLINK("C:/Users/gellison/OneDrive - FEMA/Desktop/LocalCopy/PersonalNotes/Python Resources/Programs/Root/Trunk1/Limb 11/Bough 11/Branch 1/Leaf 2.txt", \"Leaf 2.txt\")'],[f'=HYPERLINK("C:/Users/gellison/OneDrive - FEMA/Desktop/LocalCopy/PersonalNotes/Python Resources/Programs/Root/Trunk1/Limb 11/Bough 11", \"Bough 11\")',
            f'=HYPERLINK("C:/Users/gellison/OneDrive - FEMA/Desktop/LocalCopy/PersonalNotes/Python Resources/Programs/Root/Trunk1/Limb 11/Bough 11/Branch 1", \"Branch 1\")',
            f'=HYPERLINK("C:/Users/gellison/OneDrive - FEMA/Desktop/LocalCopy/PersonalNotes/Python Resources/Programs/Root/Trunk1/Limb 11/Bough 11/Branch 1/Leaf 2.txt", \"Leaf 2.txt\")']
]



with open('output.txt','r') as grove:
    forest = grove.read()
    
print(os.path.basename(r"C:/Users/gellison/OneDrive - FEMA/Desktop/LocalCopy/PersonalNotes/Python Resources/Programs/Root/Trunk1/Limb 11/Bough 11/Branch 1"))

forest = ast.literal_eval(forest) 
#metadata = forest[len(forest)-1]
#print(metadata)

def getList(dict):
    return [*dict]

def namer(dict):
    for x in dict:
        return x 

numtrunk = len(forest) - 1

#Beginning of a Parser. Break down the Trunks and Pull apart the data we've grabbed
for x in forest[0:numtrunk]:
    a = getList(forest)
    print('\n', type(a))
    for y in a:
        trunkn = os.path.basename(namer(y))
        trunknlist.append(trunkn)
        print('\n',y, print(os.path.basename(r"C:/Users/gellison/OneDrive - FEMA/Desktop/LocalCopy/PersonalNotes/Python Resources/Programs/Root/Trunk1/Limb 11/Bough 11/Branch 1")))
        


#for limb in limblist:
#    wslimb = wb.create_sheet(title=f"{limb}")
#    df = pd.DataFrame(templist, columns=['BoughFolder','BranchFolder','Branchname'])
#    writer = pd.ExcelWriter(dest_filename, engine='openpyxl')
#    writer.book = wb
#    writer.sheets = {ws.title: ws for ws in wb.worksheets}
#    for sheetname in writer.sheets:
#        df.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index=False, header=False)
#    writer.save()
#    
    
#    wslimb['A1'] = f'=HYPERLINK("Limbpath", \"{limb}\")'
    
wb.save(filename = dest_filename)

































