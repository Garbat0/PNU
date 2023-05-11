# -*- coding: utf-8 -*-
"""
Created on Wed Mar 29 17:56:46 2023

@author: gelliso1
"""

import openpyxl



def format_sheets():
    # Define the county sheets and their names
    county_sheets = ["Charlotte", "Collier", "DeSoto", "Glades", "Hendry", "Highlands", "Lee", "Lee_FMB", "Lee_CFM", "Manatee", "Okeechobee", "Sarasota"]
    county_names = ["Name", "Date", "Reporter", "Jurisdiction(s)", "Interaction Type", "RSF(s)/Advisor", "Overview", "Completion", "Due Date", "Address", "MAX-TRAX Link", "Common Drive", "B6 Teams", "Additional Notes", "Key"]
    
    # Loop through each sheet and format the columns
    for ws in wb.worksheets:
        if ws.title in county_sheets:
            ws.column_dimensions['A'].width = 3
            ws.column_dimensions['B'].width = 3
            ws['A:B'].horizontal_alignment = openpyxl.styles.Alignment(horizontal='center')
            ws['A:B'].vertical_alignment = openpyxl.styles.Alignment(vertical='top')
            ws['A:B'].font = openpyxl.styles.Font(name='Times New Roman', size=12)
            
            # Set the values of the cells in row 2 to the county names array values.
            for i in range(len(county_names)):
                ws.cell(row=2, column=i+3).value = county_names[i]
                
            # Set the value of cell C2 to Name in each sheet.
            ws['C2'] = 'Name'
            
    return wb